from datetime import datetime
from pathlib import Path
from typing import NamedTuple, List
from os import remove as os_remove
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class FileWorkers(NamedTuple):
    workbook: Workbook
    worksheet: Worksheet


class OpenpyxlWorker:
    xlsx_format = ".xlsx"
    workbook_name = "1"
    date_row = 1
    date_column = 10

    def __init__(
        self, dates: List[datetime], path_list: List[Path], save_dir: Path
    ) -> None:
        self.dates = dates
        self.path_list = path_list
        self.save_dir = save_dir

    def __read_file(self, path: Path):
        workbook = load_workbook(path, data_only=True)
        worksheet = workbook[self.workbook_name]
        return FileWorkers(workbook, worksheet)

    def __write_date(self, worksheet: Worksheet, date: str):
        worksheet.cell(self.date_row, self.date_column, date)

    def __save(self, workbook: Workbook, name: Path):
        workbook.save(name)

    def __remove_files(self, saved_paths: List[Path]):
        saved_files = tuple(Path(self.save_dir).glob(f"*{self.xlsx_format}"))
        files_to_remove = [file for file in saved_files if file not in saved_paths]

        for file in files_to_remove:
            os_remove(file)

    def generate_files(self):
        saved_paths = []

        for index, date in enumerate(self.dates):
            str_date = date.strftime("%d.%m.%Y")
            save_path = Path(self.save_dir, f"{str_date}{self.xlsx_format}")

            file_workers = self.__read_file(self.path_list[index])
            self.__write_date(file_workers.worksheet, str_date)
            self.__save(file_workers.workbook, save_path)
            saved_paths.append(save_path)

        self.__remove_files(saved_paths)
